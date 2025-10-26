import csv
from datetime import time, timedelta, datetime
from decimal import Decimal
import io
from openpyxl.styles import Font
from typing import Dict, List, Optional

from fastapi import FastAPI,  APIRouter

from fastapi.params import Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import extract, func, literal, or_, union_all
from sqlalchemy.orm import Session, aliased
from starlette import status

from starlette.exceptions import HTTPException

from models.AuditTrail import AuditEntityEnum
from models.BatchStock import BatchStock, FifoLog
from models.InventoryLedger import InventoryLedger, SourceTypeEnum
from models.KodeLambung import KodeLambung
from models.Customer import Customer
from models.Item import Item
from models.Pembelian import Pembelian, PembelianItem, StatusPembelianEnum
from models.Penjualan import Penjualan, PenjualanItem
from models.Vendor import Vendor
from schemas.PaginatedResponseSchemas import PaginatedResponse
from database import  get_db
from schemas.UtilsSchemas import DashboardStatistics, ItemStockAdjustmentReportRow, LabaRugiDetailRow, LabaRugiResponse, PurchaseReportResponse, PurchaseReportRow, \
    SalesReportRow, SalesReportResponse, SalesTrendResponse, SalesTrendDataPoint, StockAdjustmentReportResponse, StockAdjustmentReportRow
from services.audit_services import AuditService

router =APIRouter()


def get_status(this_month: float, last_month: float) -> str:
    if this_month > last_month:
        return "up"
    elif this_month < last_month:
        return "down"
    else:
        return "neutral"

@router.get("/statistics", status_code=status.HTTP_200_OK, response_model=DashboardStatistics)
async def get_dashboard_statistics(db: Session = Depends(get_db)):
    now = datetime.now()
    this_month = now.month
    this_year = now.year
    
    # Better way to calculate last month
    first_day_this_month = now.replace(day=1)
    last_month_date = first_day_this_month - timedelta(days=1)
    last_month = last_month_date.month
    last_month_year = last_month_date.year

    # Helper function to safely calculate percentage
    def calculate_percentage(current, previous):
        if previous == 0:
            return 100.0 if current > 0 else 0.0
        return ((current - previous) / previous * 100)

    # Products
    total_products = db.query(Item).filter(Item.is_active == True).count()
    this_month_products = db.query(Item).filter(
        extract('month', Item.created_at) == this_month,
        extract('year', Item.created_at) == this_year
    ).count()
    last_month_products = db.query(Item).filter(
        extract('month', Item.created_at) == last_month,
        extract('year', Item.created_at) == last_month_year
    ).count()
    percentage_month_products = calculate_percentage(this_month_products, last_month_products)
    status_month_products = get_status(this_month_products, last_month_products)

    # Customers
    total_customer = db.query(Customer).filter(Customer.is_active == True).count()
    this_month_customer = db.query(Customer).filter(
        extract('month', Customer.created_at) == this_month,
        extract('year', Customer.created_at) == this_year
    ).count()
    last_month_customer = db.query(Customer).filter(
        extract('month', Customer.created_at) == last_month,
        extract('year', Customer.created_at) == last_month_year
    ).count()
    percentage_month_customer = calculate_percentage(this_month_customer, last_month_customer)
    status_month_customer = get_status(this_month_customer, last_month_customer)

    # Pembelian - More robust null handling
    total_pembelian = db.query(func.coalesce(func.sum(Pembelian.total_price), 0)).scalar() or 0
    this_month_pembelian = db.query(func.coalesce(func.sum(Pembelian.total_price), 0)).filter(
        extract('month', Pembelian.created_at) == this_month,
        extract('year', Pembelian.created_at) == this_year
    ).scalar() or 0
    last_month_pembelian = db.query(func.coalesce(func.sum(Pembelian.total_price), 0)).filter(
        extract('month', Pembelian.created_at) == last_month,
        extract('year', Pembelian.created_at) == last_month_year
    ).scalar() or 0
    percentage_month_pembelian = calculate_percentage(this_month_pembelian, last_month_pembelian)
    status_month_pembelian = get_status(this_month_pembelian, last_month_pembelian)

    # Penjualan - More robust null handling
    total_penjualan = db.query(func.coalesce(func.sum(Penjualan.total_price), 0)).scalar() or 0
    this_month_penjualan = db.query(func.coalesce(func.sum(Penjualan.total_price), 0)).filter(
        extract('month', Penjualan.created_at) == this_month,
        extract('year', Penjualan.created_at) == this_year
    ).scalar() or 0
    last_month_penjualan = db.query(func.coalesce(func.sum(Penjualan.total_price), 0)).filter(
        extract('month', Penjualan.created_at) == last_month,
        extract('year', Penjualan.created_at) == last_month_year
    ).scalar() or 0
    percentage_month_penjualan = calculate_percentage(this_month_penjualan, last_month_penjualan)
    status_month_penjualan = get_status(this_month_penjualan, last_month_penjualan)

    return DashboardStatistics(
        total_products=total_products,
        percentage_month_products=percentage_month_products,
        status_month_products=status_month_products,

        total_customer=total_customer,
        percentage_month_customer=percentage_month_customer,
        status_month_customer=status_month_customer,

        total_pembelian=total_pembelian,
        percentage_month_pembelian=percentage_month_pembelian,
        status_month_pembelian=status_month_pembelian,

        total_penjualan=total_penjualan,
        percentage_month_penjualan=percentage_month_penjualan,
        status_month_penjualan=status_month_penjualan
    )
    
    
@router.get("/laba-rugi", status_code=status.HTTP_200_OK, response_model=LabaRugiResponse)
async def get_laba_rugi(
    from_date: datetime = Query(..., description="Start datetime (ISO-8601)"),
    to_date: Optional[datetime] = Query(None, description="End datetime (inclusive)"),
    item_id: Optional[int] = Query(None, description="Filter by specific item"),
    skip: int = Query(0, ge=0, description="Number of records to skip"),
    limit: int = Query(100, ge=1, le=1000, description="Maximum number of records"),
    db: Session = Depends(get_db),
):
    """
    Get Laba Rugi (Profit & Loss) report based on FIFO logs.
    Shows detailed breakdown by invoice with HPP calculation.
    """
    if to_date is None:
        to_date = datetime.now()

    from_date_only = from_date.date()
    to_date_only = to_date.date()

    # Query FifoLog with item details
    query = (
        db.query(
            FifoLog.invoice_date,
            FifoLog.invoice_id,
            FifoLog.item_id,
            Item.code.label("item_code"),
            Item.name.label("item_name"),
            func.sum(FifoLog.qty_terpakai).label("qty_terjual"),
            func.sum(FifoLog.total_hpp).label("total_hpp"),
            func.sum(FifoLog.total_penjualan).label("total_penjualan"),
            func.sum(FifoLog.laba_kotor).label("laba_kotor"),
            FifoLog.harga_jual,  # Assume same price per invoice
        )
        .join(Item, Item.id == FifoLog.item_id)
        .filter(
            FifoLog.invoice_date >= from_date_only,
            FifoLog.invoice_date <= to_date_only,
        )
        .group_by(
            FifoLog.invoice_date,
            FifoLog.invoice_id,
            FifoLog.item_id,
            Item.code,
            Item.name,
            FifoLog.harga_jual,
        )
        .order_by(FifoLog.invoice_date.asc(), FifoLog.invoice_id.asc())
    )

    # Apply optional item filter
    if item_id is not None:
        query = query.filter(FifoLog.item_id == item_id)

    # Get total count before pagination
    total_count = query.count()

    # Apply pagination
    results = query.offset(skip).limit(limit).all()

    # Format response
    detail_rows = []
    grand_total_hpp = Decimal("0")
    grand_total_penjualan = Decimal("0")
    grand_total_laba = Decimal("0")
    total_qty = 0

    for row in results:
        qty = row.qty_terjual or 0
        total_hpp = row.total_hpp or Decimal("0")
        total_penjualan = row.total_penjualan or Decimal("0")
        laba_kotor = row.laba_kotor or Decimal("0")
        
        # Calculate HPP per unit
        hpp_per_unit = total_hpp / qty if qty > 0 else Decimal("0")

        detail_rows.append(LabaRugiDetailRow(
            tanggal=datetime.combine(row.invoice_date, datetime.min.time()),
            no_invoice=row.invoice_id,
            item_code=row.item_code or "N/A",
            item_name=row.item_name or "N/A",
            qty_terjual=qty,
            hpp=hpp_per_unit,
            total_hpp=total_hpp,
            harga_jual=row.harga_jual or Decimal("0"),
            total_penjualan=total_penjualan,
            laba_kotor=laba_kotor,
        ))

        # Accumulate totals
        grand_total_hpp += total_hpp
        grand_total_penjualan += total_penjualan
        grand_total_laba += laba_kotor
        total_qty += qty

    title = f"Laporan Laba Rugi {from_date:%d/%m/%Y} - {to_date:%d/%m/%Y}"

    return LabaRugiResponse(
        title=title,
        date_from=from_date,
        date_to=to_date,
        details=detail_rows,
        total_qty=total_qty,
        total_hpp=grand_total_hpp,
        total_penjualan=grand_total_penjualan,
        total_laba_kotor=grand_total_laba,
        total=total_count,
    )


@router.get(
    "/laba-rugi/download",
    status_code=status.HTTP_200_OK,
    summary="Download Laporan Laba Rugi as XLSX",
)
async def download_laba_rugi(
    from_date: datetime = Query(..., description="Start datetime (ISO-8601)"),
    to_date: Optional[datetime] = Query(None, description="End datetime (inclusive)"),
    item_id: Optional[int] = Query(None, description="Filter by specific item"),
    db: Session = Depends(get_db),
):
    """
    Download profit and loss report (Laba Rugi) as XLSX file with FIFO detail.
    Shows detailed breakdown by invoice with HPP calculation per batch.
    """
    if to_date is None:
        to_date = datetime.now()

    from_date_only = from_date.date()
    to_date_only = to_date.date()

    # Query FifoLog with item details
    query = (
        db.query(
            FifoLog.invoice_date,
            FifoLog.invoice_id,
            FifoLog.item_id,
            Item.code.label("item_code"),
            Item.name.label("item_name"),
            func.sum(FifoLog.qty_terpakai).label("qty_terjual"),
            func.sum(FifoLog.total_hpp).label("total_hpp"),
            func.sum(FifoLog.total_penjualan).label("total_penjualan"),
            func.sum(FifoLog.laba_kotor).label("laba_kotor"),
            FifoLog.harga_jual,
        )
        .join(Item, Item.id == FifoLog.item_id)
        .filter(
            FifoLog.invoice_date >= from_date_only,
            FifoLog.invoice_date <= to_date_only,
        )
        .group_by(
            FifoLog.invoice_date,
            FifoLog.invoice_id,
            FifoLog.item_id,
            Item.code,
            Item.name,
            FifoLog.harga_jual,
        )
        .order_by(FifoLog.invoice_date.asc(), FifoLog.invoice_id.asc())
    )

    if item_id is not None:
        query = query.filter(FifoLog.item_id == item_id)

    results = query.all()

    # Get detailed batch usage for notes section
    fifo_logs_query = (
        db.query(FifoLog, BatchStock, Item)
        .join(BatchStock, BatchStock.id_batch == FifoLog.id_batch)
        .join(Item, Item.id == FifoLog.item_id)
        .filter(
            FifoLog.invoice_date >= from_date_only,
            FifoLog.invoice_date <= to_date_only,
        )
        .order_by(FifoLog.invoice_date.asc(), FifoLog.invoice_id.asc())
    )

    if item_id is not None:
        fifo_logs_query = fifo_logs_query.filter(FifoLog.item_id == item_id)

    fifo_logs = fifo_logs_query.all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laba Rugi"

    # Title
    ws.append([f"Laporan Laba Rugi"])
    ws.append([f"Periode: {from_date:%d/%m/%Y} - {to_date:%d/%m/%Y}"])
    ws.append([])

    # Headers
    headers = [
        "Tanggal",
        "No. Invoice",
        "Item Code",
        "Item",
        "Qty Terjual",
        "HPP (per unit)",
        "Total HPP",
        "Harga Jual (per unit)",
        "Total Penjualan",
        "Laba Kotor",
    ]
    ws.append(headers)

    # Make headers bold
    for cell in ws[4]:
        cell.font = Font(bold=True)

    # Data rows
    grand_total_qty = 0
    grand_total_hpp = Decimal("0")
    grand_total_penjualan = Decimal("0")
    grand_total_laba = Decimal("0")

    for row in results:
        qty = row.qty_terjual or 0
        total_hpp = row.total_hpp or Decimal("0")
        total_penjualan = row.total_penjualan or Decimal("0")
        laba_kotor = row.laba_kotor or Decimal("0")
        
        hpp_per_unit = total_hpp / qty if qty > 0 else Decimal("0")

        ws.append([
            row.invoice_date.strftime("%d/%m/%Y"),
            row.invoice_id,
            row.item_code or "N/A",
            row.item_name or "N/A",
            int(qty),
            float(hpp_per_unit),
            float(total_hpp),
            float(row.harga_jual or 0),
            float(total_penjualan),
            float(laba_kotor),
        ])

        grand_total_qty += qty
        grand_total_hpp += total_hpp
        grand_total_penjualan += total_penjualan
        grand_total_laba += laba_kotor

    # Grand total row
    ws.append([
        "TOTAL",
        "",
        "",
        "",
        int(grand_total_qty),
        "",
        float(grand_total_hpp),
        "",
        float(grand_total_penjualan),
        float(grand_total_laba),
    ])
    
    # Make total row bold
    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = Font(bold=True)

    # Add notes section with detailed batch calculations
    ws.append([])
    ws.append([])
    ws.append(["Notes:"])
    ws[ws.max_row]["A"].font = Font(bold=True)
    ws.append([])

    # Group fifo logs by invoice
    invoice_batches = {}
    for fifo_log, batch, item in fifo_logs:
        key = (fifo_log.invoice_date, fifo_log.invoice_id, item.name)
        if key not in invoice_batches:
            invoice_batches[key] = []
        invoice_batches[key].append({
            'batch_id': batch.id_batch,
            'qty': fifo_log.qty_terpakai,
            'harga_beli': batch.harga_beli,
            'hpp': fifo_log.total_hpp,
        })

    # Write batch detail notes
    for (inv_date, inv_id, item_name), batches in invoice_batches.items():
        ws.append([f"Perhitungan HPP pada tanggal {inv_date.strftime('%d/%m/%Y')} - {inv_id} ({item_name}):"])
        ws[ws.max_row]["A"].font = Font(bold=True)
        
        # Build formula string
        formula_parts = []
        for b in batches:
            formula_parts.append(f"(Batch-{b['batch_id']}: {b['qty']}qty × Rp {b['harga_beli']:,.2f})")
        
        formula_str = " + ".join(formula_parts)
        total_hpp = sum(b['hpp'] for b in batches)
        
        ws.append([f"Rumus FIFO = {formula_str}"])
        ws.append([f"Total HPP = Rp {total_hpp:,.2f}"])
        ws.append([])

    # Format columns
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 15

    # Apply number formatting for currency columns
    for row in ws.iter_rows(min_row=5, max_row=total_row):
        # HPP per unit (F)
        if row[5].value and isinstance(row[5].value, (int, float)):
            row[5].number_format = '#,##0.00'
        # Total HPP (G)
        if row[6].value and isinstance(row[6].value, (int, float)):
            row[6].number_format = '#,##0.00'
        # Harga Jual (H)
        if row[7].value and isinstance(row[7].value, (int, float)):
            row[7].number_format = '#,##0.00'
        # Total Penjualan (I)
        if row[8].value and isinstance(row[8].value, (int, float)):
            row[8].number_format = '#,##0.00'
        # Laba Kotor (J)
        if row[9].value and isinstance(row[9].value, (int, float)):
            row[9].number_format = '#,##0.00'

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"laba_rugi_{from_date:%Y%m%d}_{to_date:%Y%m%d}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )

@router.get(
    "/penjualan",
    status_code=status.HTTP_200_OK,
    response_model=PaginatedResponse[SalesReportRow],
    summary="Laporan Penjualan (consolidated per sale)",
)

async def get_penjualan_laporan(
        from_date: datetime = Query(..., description="Start datetime (inclusive)"),
        to_date: Optional[datetime] = Query(None, description="End datetime (inclusive)"),
        customer_id: Optional[int] = Query(None, description="Customer ID"),
        kode_lambung_id: Optional[int] = Query(None, description="Kode Lambung ID"),
        skip: int = Query(0, ge=0, description="Number of records to skip"),
        limit: int = Query(100, ge=1, le=1000, description="Maximum number of records to return"),
        db: Session = Depends(get_db),
):
    """
    Returns one row per sale with concatenated item details and aggregated totals.
    Shows penjualan's kode_lambung only (Customer doesn't have a kode_lambung FK).
    """

    if to_date is None:
        to_date = datetime.now()

    sales_query = (
        db.query(
            Penjualan.id,
            Penjualan.sales_date.label("date"),
             Penjualan.sales_due_date,
            Penjualan.customer_name,
            Customer.name.label("customer_name_rel"),
            Penjualan.kode_lambung_id,
            KodeLambung.name.label("penjualan_kode_lambung"),
            Penjualan.no_penjualan,
            Penjualan.status_pembayaran,
        )
        .join(Customer, Customer.id == Penjualan.customer_id, isouter=True)
        .join(KodeLambung, KodeLambung.id == Penjualan.kode_lambung_id, isouter=True)
        .filter(
            Penjualan.is_deleted == False,
            Penjualan.status_penjualan != StatusPembelianEnum.DRAFT,
            Penjualan.sales_date >= from_date,
            Penjualan.sales_date <= to_date,
            )
    )

    if customer_id is not None:
        sales_query = sales_query.filter(Penjualan.customer_id == customer_id)

    if kode_lambung_id is not None:
        sales_query = sales_query.filter(Penjualan.kode_lambung_id == kode_lambung_id)

    sales_query = sales_query.order_by(Penjualan.sales_date.asc(), Penjualan.no_penjualan.asc())

    total_count = sales_query.count()
    sales = sales_query.offset(skip).limit(limit).all()

    def _dec(x) -> Decimal:
        return Decimal(str(x or 0))

    report_rows: List[SalesReportRow] = []

    for sale in sales:
        items = (
            db.query(
                PenjualanItem.qty,
                PenjualanItem.unit_price,
                PenjualanItem.discount,
                PenjualanItem.tax_percentage,
                Item.code.label("item_code"),
                Item.name.label("item_name"),
            )
            .join(Item, Item.id == PenjualanItem.item_id, isouter=True)
            .filter(PenjualanItem.penjualan_id == sale.id)
            .all()
        )

        item_codes, item_names = [], []
        total_subtotal = total_discount = total_tax = Decimal("0")
        total_qty = 0

        for it in items:
            qty = int(it.qty or 0)
            price = _dec(it.unit_price)
            discount = _dec(it.discount)
            tax_pct = Decimal(str(it.tax_percentage or 0))

            subtotal = price * qty
            total_after_discount = max(subtotal - discount, Decimal("0"))
            tax = total_after_discount * tax_pct / Decimal(100)

            item_codes.append(it.item_code or "N/A")
            item_names.append(it.item_name or "N/A")

            total_subtotal += subtotal
            total_discount += discount
            total_tax += tax
            total_qty += qty

        final_total = max(total_subtotal - total_discount, Decimal("0"))
        grand_total = final_total + total_tax

        report_rows.append(
            SalesReportRow(
                date=sale.date,
                sales_due_date=sale.sales_due_date,
                customer=sale.customer_name or sale.customer_name_rel or "—",
                kode_lambung_rel=sale.penjualan_kode_lambung,
                kode_lambung_penjualan=sale.penjualan_kode_lambung,
                no_penjualan=sale.no_penjualan,
                status=(sale.status_pembayaran.name.capitalize()
                        if hasattr(sale.status_pembayaran, "name")
                        else str(sale.status_pembayaran)),
                item_code=", ".join(item_codes) or "No items",
                item_name=", ".join(item_names) or "No items",
                qty=total_qty,
                price=(total_subtotal / total_qty) if total_qty > 0 else Decimal("0"),
                sub_total=total_subtotal,
                total=final_total,
                tax=total_tax,
                grand_total=grand_total,
            )
        )

    title = f"Laporan Penjualan {from_date:%d/%m/%Y} - {to_date:%d/%m/%Y}"
    return SalesReportResponse(
        title=title,
        date_from=from_date,
        date_to=to_date,
        data=report_rows,
        total=total_count,
    )
@router.get(
    "/penjualan/download",
    status_code=status.HTTP_200_OK,
    summary="Download Laporan Penjualan as CSV (all records)",
)


@router.get(
    "/penjualan/download",
    status_code=status.HTTP_200_OK,
    summary="Download Laporan Penjualan as XLSX (all records)",
)
async def download_penjualan_laporan(
    from_date: datetime = Query(..., description="Start datetime (inclusive)"),
    to_date: Optional[datetime ] = Query(None, description="End datetime (inclusive)"),
    customer_id: Optional[int] = Query(None, description="Customer ID"),
    kode_lambung_id: Optional[int ] = Query(None, description="Kode Lambung ID"),
    db: Session = Depends(get_db),
):
    """
    Download complete sales report as XLSX without pagination.
    Returns all records matching the date filter — one row per sale with concatenated items.
    """

    if to_date is None:
        to_date = datetime.now()

    # Base query (same as before)
    sales_query = (
        db.query(
            Penjualan.id,
            Penjualan.sales_date.label("date"),
            Penjualan.sales_due_date,
            Penjualan.customer_name,
            Customer.name.label("customer_name_rel"),
            Penjualan.kode_lambung_id,
            KodeLambung.name.label("penjualan_kode_lambung"),
            Penjualan.no_penjualan,
            Penjualan.status_pembayaran,
        )
        .join(Customer, Customer.id == Penjualan.customer_id, isouter=True)
        .join(KodeLambung, KodeLambung.id == Penjualan.kode_lambung_id, isouter=True)
        .filter(
            Penjualan.is_deleted.is_(False),
            Penjualan.status_penjualan != StatusPembelianEnum.DRAFT,
            Penjualan.sales_date >= from_date,
            Penjualan.sales_date <= to_date,
        )
    )

    # Optional filters
    if customer_id is not None:
        sales_query = sales_query.filter(Penjualan.customer_id == customer_id)
    if kode_lambung_id is not None:
        sales_query = sales_query.filter(Penjualan.kode_lambung_id == kode_lambung_id)

    sales_query = sales_query.order_by(Penjualan.sales_date.asc(), Penjualan.no_penjualan.asc())
    sales = sales_query.all()

    def _dec(x) -> Decimal:
        return Decimal(str(x or 0))

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Penjualan"

    # Write header
    headers = [
        "Date",
        "Due Date",
        "Customer",
        "Kode Lambung",
        "No Penjualan",
        "Status",
        "Item Code",
        "Item Name",
        "Qty",
        "Price",
        "Sub Total",
        "Total",
        "Tax",
        "Grand Total",
    ]
    ws.append(headers)


    # Write rows
    for sale in sales:
        items_query = (
            db.query(
                PenjualanItem.qty,
                PenjualanItem.unit_price,
                PenjualanItem.discount,
                PenjualanItem.tax_percentage,
                Item.code.label("item_code"),
                Item.name.label("item_name"),
            )
            .join(Item, Item.id == PenjualanItem.item_id, isouter=True)
            .filter(PenjualanItem.penjualan_id == sale.id)
            .all()
        )

        item_codes, item_names = [], []
        total_subtotal, total_discount, total_tax = Decimal("0"), Decimal("0"), Decimal("0")
        total_qty = 0

        for item in items_query:
            item_code = item.item_code or "N/A"
            item_name = item.item_name or "N/A"
            qty = int(item.qty or 0)
            price = _dec(item.unit_price)
            item_subtotal = price * qty
            item_discount = _dec(item.discount)
            item_total = max(item_subtotal - item_discount, Decimal("0"))
            tax_pct = Decimal(str(item.tax_percentage or 0))
            item_tax = item_total * tax_pct / Decimal(100)

            item_codes.append(item_code)
            item_names.append(item_name)

            total_subtotal += item_subtotal
            total_discount += item_discount
            total_tax += item_tax
            total_qty += qty

        # Final totals
        final_total = max(total_subtotal - total_discount, Decimal("0"))
        grand_total = final_total + total_tax
        item_codes_str = ", ".join(item_codes) if item_codes else "No items"
        item_names_str = ", ".join(item_names) if item_names else "No items"

        customer_name = sale.customer_name or sale.customer_name_rel or "—"
        kode_lambung = sale.penjualan_kode_lambung or ""

        ws.append([
            sale.date.strftime("%d/%m/%Y") if sale.date else "",
            sale.sales_due_date.strftime("%d/%m/%Y") if sale.sales_due_date else "",
            customer_name,
            kode_lambung,
            sale.no_penjualan or "",
            (sale.status_pembayaran.name.capitalize() if hasattr(sale.status_pembayaran, "name")
             else str(sale.status_pembayaran)),
            item_codes_str,
            item_names_str,
            total_qty,
            float(total_subtotal / total_qty) if total_qty > 0 else 0.0,
            float(total_subtotal),
            float(final_total),
            float(total_tax),
            float(grand_total),
        ])

    # Save workbook to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"laporan_penjualan_{from_date:%Y%m%d}_{to_date:%Y%m%d}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )
@router.get("/pembelian")
async def get_pembelian_laporan(
    from_date: datetime = Query(...),
    to_date: Optional[datetime] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: Session = Depends(get_db),
):
    """
    Returns one row per purchase with concatenated item details and aggregated totals.
    """
    if to_date is None:
        to_date = datetime.now()

    # Base query to get purchases
    purchases_query = (
        db.query(
            Pembelian.id,
            Pembelian.sales_date.label("date"),
            Pembelian.sales_due_date,
            Vendor.name.label("vendor_name_rel"),
            Pembelian.no_pembelian,
            Pembelian.status_pembayaran,
        )
        .join(Vendor, Vendor.id == Pembelian.vendor_id, isouter=True)
        .filter(
            Pembelian.is_deleted.is_(False),
            Pembelian.status_pembelian != StatusPembelianEnum.DRAFT,
            Pembelian.sales_date >= from_date.date(),
            Pembelian.sales_date <= to_date.date(),
        )
        .order_by(Pembelian.sales_date.asc(), Pembelian.no_pembelian.asc())
    )

    # Get total count
    total_count = purchases_query.count()

    # Get paginated purchases
    purchases = purchases_query.offset(skip).limit(limit).all()

    def _dec(x) -> Decimal:
        return Decimal(str(x or 0))

    report_rows: List[PurchaseReportRow] = []
    
    for purchase in purchases:
        # Get all items for this purchase with proper Item.name
        items_query = (
            db.query(
                Item.sku.label("item_sku"),
                Item.name.label("item_name"),
                Item.code.label("item_code"),
            
                PembelianItem.qty,
                PembelianItem.unit_price,
                PembelianItem.discount,
                PembelianItem.tax_percentage,
            )
            .join(Item, Item.id == PembelianItem.item_id, isouter=True)
            .filter(PembelianItem.pembelian_id == purchase.id)
            .all()
        )

        # Concatenate item details and calculate totals
        item_codes = []
        item_names = []  # Separate list for item names
        total_subtotal = Decimal("0")
        total_discount = Decimal("0")
        total_tax = Decimal("0")
        total_qty = 0

        for item in items_query:
            # Build item detail strings
            item_code = item.item_code or item.item_sku or "N/A"
            # Use Item.name from the joined table
            item_name = item.item_name or "N/A"
            qty = int(item.qty or 0)
            
            item_codes.append(item_code)
            item_names.append(item_name)
            
            # Calculate totals
            price = _dec(item.unit_price)
            item_subtotal = price * qty
            item_discount = _dec(item.discount)
            item_total = item_subtotal - item_discount
            if item_total < 0:
                item_total = Decimal("0")
            
            tax_pct = Decimal(str(item.tax_percentage or 0))
            item_tax = (item_total * tax_pct / Decimal(100))
            
            total_subtotal += item_subtotal
            total_discount += item_discount
            total_tax += item_tax
            total_qty += qty

        # Join items with comma
        item_codes_str = ", ".join(item_codes) if item_codes else "No items"
        item_names_str = ", ".join(item_names) if item_names else "No items"
        
        # Calculate final totals
        final_total = total_subtotal - total_discount
        if final_total < 0:
            final_total = Decimal("0")
        grand_total = final_total + total_tax

        report_rows.append(
            PurchaseReportRow(
                date=purchase.date,
                sales_due_date = purchase.sales_due_date,
                vendor=purchase.vendor_name_rel or "—",
                no_pembelian=purchase.no_pembelian,
                status=(purchase.status_pembayaran.name.capitalize() if hasattr(purchase.status_pembayaran, "name")
                        else str(purchase.status_pembayaran)),
                item_code=item_codes_str,  # Concatenated item codes
                item_name=item_names_str,  # Concatenated actual item names from Item.name
                qty=total_qty,
                price=total_subtotal / total_qty if total_qty > 0 else Decimal("0"),  # Average price
                sub_total=total_subtotal,
                total=final_total,
                tax=total_tax,
                grand_total=grand_total,
            )
        )

    title = f"Laporan Pembelian {from_date:%d/%m/%Y} - {to_date:%d/%m/%Y}"
    return PurchaseReportResponse(
        title=title,
        date_from=from_date,
        date_to=to_date,
        data=report_rows,
        total=total_count,
    )

@router.get(
    "/pembelian/download",
    status_code=status.HTTP_200_OK,
    summary="Download Laporan Pembelian as XLSX (all records)",
)
async def download_pembelian_laporan(
    from_date: datetime = Query(..., description="Start datetime (inclusive)"),
    to_date: Optional[datetime ] = Query(None, description="End datetime (inclusive)"),
    db: Session = Depends(get_db),
):
    if to_date is None:
        to_date = datetime.now()

    purchases_query = (
        db.query(
            Pembelian.id,
            Pembelian.sales_date.label("date"),
            Pembelian.sales_due_date,
            Vendor.name.label("vendor_name_rel"),
            Pembelian.no_pembelian,
            Pembelian.status_pembayaran,
        )
        .join(Vendor, Vendor.id == Pembelian.vendor_id, isouter=True)
        .filter(
            Pembelian.is_deleted.is_(False),
            Pembelian.status_pembelian != StatusPembelianEnum.DRAFT,
            Pembelian.sales_date >= from_date.date(),
            Pembelian.sales_date <= to_date.date(),
        )
        .order_by(Pembelian.sales_date.asc(), Pembelian.no_pembelian.asc())
    )

    purchases = purchases_query.all()

    def _dec(x) -> Decimal:
        return Decimal(str(x or 0))

    # Create Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Pembelian"

    # Write header
    ws.append([
        'Date', 'Due Date', 'Vendor', 'No Pembelian', 'Status',
        'Item Code', 'Item Name', 'Qty', 'Price', 'Sub Total',
        'Total', 'Tax', 'Grand Total'
    ])

    for purchase in purchases:
        items_query = (
            db.query(
                Item.sku.label("item_sku"),
                Item.name.label("item_name"),
                Item.code.label("item_code"),
                PembelianItem.qty,
                PembelianItem.unit_price,
                PembelianItem.discount,
                PembelianItem.tax_percentage,
            )
            .join(Item, Item.id == PembelianItem.item_id, isouter=True)
            .filter(PembelianItem.pembelian_id == purchase.id)
            .all()
        )

        item_codes, item_names = [], []
        total_subtotal, total_discount, total_tax = Decimal("0"), Decimal("0"), Decimal("0")
        total_qty = 0

        for item in items_query:
            item_code = item.item_code or item.item_sku or "N/A"
            item_name = item.item_name or "N/A"
            qty = int(item.qty or 0)

            item_codes.append(item_code)
            item_names.append(item_name)

            price = _dec(item.unit_price)
            item_subtotal = price * qty
            item_discount = _dec(item.discount)
            item_total = max(item_subtotal - item_discount, Decimal("0"))
            tax_pct = Decimal(str(item.tax_percentage or 0))
            item_tax = item_total * tax_pct / Decimal(100)

            total_subtotal += item_subtotal
            total_discount += item_discount
            total_tax += item_tax
            total_qty += qty

        item_codes_str = ", ".join(item_codes) if item_codes else "No items"
        item_names_str = ", ".join(item_names) if item_names else "No items"
        final_total = max(total_subtotal - total_discount, Decimal("0"))
        grand_total = final_total + total_tax

        ws.append([
            purchase.date.strftime('%d/%m/%Y') if purchase.date else '',
            purchase.sales_due_date.strftime('%d/%m/%Y') if purchase.sales_due_date else '',
            purchase.vendor_name_rel or "—",
            purchase.no_pembelian or '',
            (purchase.status_pembayaran.name.capitalize() if hasattr(purchase.status_pembayaran, "name")
             else str(purchase.status_pembayaran)),
            item_codes_str,
            item_names_str,
            total_qty,
            float(total_subtotal / total_qty) if total_qty > 0 else 0.0,
            float(total_subtotal),
            float(final_total),
            float(total_tax),
            float(grand_total),
        ])

    # Save workbook to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"laporan_pembelian_{from_date:%Y%m%d}_{to_date:%Y%m%d}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )
@router.get("/tren-penjualan", response_model=SalesTrendResponse)
async def get_sales_trend(
        period: str = Query(
            "mtd",
            regex="^(daily|mtd|custom)$",
            description="Period: 'daily' (last 30 days), 'mtd' (month to date), 'custom' (requires from_date and to_date)"
        ),
        from_date: Optional[datetime] = Query(None, description="Start date for custom period"),
        to_date: Optional[datetime] = Query(None, description="End date for custom period"),
        db: Session = Depends(get_db)
):
    """
    Get sales trend data showing daily order count and revenue.

    - **daily**: Last 30 days from today
    - **mtd**: Month to date (from 1st of current month to today)
    - **custom**: Custom date range (requires from_date and to_date)
    """

    now = datetime.now()

    # Determine date range based on period
    if period == "daily":
        start_date = (now - timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
        title = "Tren Penjualan - 30 Hari Terakhir"

    elif period == "mtd":
        start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
        title = f"Tren Penjualan - Month to Date ({start_date.strftime('%B %Y')})"

    elif period == "custom":
        if not from_date or not to_date:
            raise HTTPException(
                status_code=400,
                detail="from_date and to_date are required for custom period"
            )
        start_date = from_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = to_date.replace(hour=23, minute=59, second=59, microsecond=999999)
        title = f"Tren Penjualan - {start_date.strftime('%d/%m/%Y')} s/d {end_date.strftime('%d/%m/%Y')}"

    # Query to get daily sales data
    daily_sales = (
        db.query(
            func.date(Penjualan.sales_date).label('sale_date'),
            func.count(Penjualan.id).label('order_count'),
            func.coalesce(func.sum(Penjualan.total_price), 0).label('revenue')
        )
        .filter(
            Penjualan.is_deleted == False,
            Penjualan.status_penjualan != StatusPembelianEnum.DRAFT,
            Penjualan.sales_date >= start_date,
            Penjualan.sales_date <= end_date,
            Penjualan.sales_date.isnot(None)  # Exclude null dates
        )
        .group_by(func.date(Penjualan.sales_date))
        .order_by(func.date(Penjualan.sales_date))
        .all()
    )

    # Convert to response format
    trend_data = []
    total_orders = 0
    total_revenue = Decimal('0')

    for sale_data in daily_sales:
        # Convert date to datetime for consistent response
        sale_datetime = datetime.combine(sale_data.sale_date, datetime.min.time())
        revenue = Decimal(str(sale_data.revenue or 0))

        trend_data.append(SalesTrendDataPoint(
            date=sale_datetime,
            order_count=sale_data.order_count,
            revenue=revenue
        ))

        total_orders += sale_data.order_count
        total_revenue += revenue

    # Fill missing dates with zero values (optional - for complete timeline)
    if period in ["daily", "mtd"]:
        # Create a complete date range
        current_date = start_date.date()
        end_date_only = end_date.date()
        existing_dates = {dp.date.date() for dp in trend_data}

        complete_data = []
        while current_date <= end_date_only:
            if current_date in existing_dates:
                # Find existing data point
                existing_point = next(dp for dp in trend_data if dp.date.date() == current_date)
                complete_data.append(existing_point)
            else:
                # Add zero data point for missing dates
                complete_data.append(SalesTrendDataPoint(
                    date=datetime.combine(current_date, datetime.min.time()),
                    order_count=0,
                    revenue=Decimal('0')
                ))
            current_date += timedelta(days=1)

        trend_data = complete_data

    return SalesTrendResponse(
        title=title,
        period=period,
        data=trend_data,
        total_orders=total_orders,
        total_revenue=total_revenue
    )



def calculate_hpp(prev_balance: Decimal, prev_hpp: Decimal, qty_in: int, price_in: Decimal) -> Decimal:
    """
    Calculate HPP using weighted average formula: ((prev_balance * prev_hpp) + (qty_in * price_in)) / (prev_balance + qty_in)
    """
    if prev_balance + qty_in == 0:
        return Decimal("0")
    
    total_value = (prev_balance * prev_hpp) + (qty_in * price_in)
    total_qty = prev_balance + qty_in
    
    return total_value / total_qty

# --- STOCK ADJUSTMENT (BatchStock + FifoLog) ---------------------------------
from datetime import datetime, date, time, timedelta
from decimal import Decimal
from typing import Dict, List, Optional, Tuple
from fastapi import APIRouter, Depends, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_
from openpyxl import Workbook
import io

router = APIRouter()

def _dt_bounds(from_date: datetime, to_date: Optional[datetime]) -> tuple[datetime, datetime, datetime]:
    """Inclusive start, exclusive end. Returns (start_dt, end_dt_excl, effective_to)."""
    if to_date is None:
        to_date = datetime.now()
    start_dt = datetime.combine(from_date.date(), time.min)
    end_dt_excl = datetime.combine(to_date.date() + timedelta(days=1), time.min)
    return start_dt, end_dt_excl, to_date

def _D(x) -> Decimal:
    if isinstance(x, Decimal): return x
    if x is None: return Decimal("0")
    return Decimal(str(x))

# ---------- JSON endpoint

@router.get(
    "/stock-adjustment",
    status_code=status.HTTP_200_OK,
    response_model=StockAdjustmentReportResponse,
)
async def get_stock_adjustment_report(
    from_date: datetime = Query(..., description="Start datetime (inclusive)"),
    to_date: Optional[datetime] = Query(None, description="End datetime (inclusive)"),
    item_id: Optional[int] = Query(None, description="Filter by specific item"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: Session = Depends(get_db),
):
    """
    Build stock adjustment from BatchStock (IN) + FifoLog (OUT).
    Shows per-item merged movements with running balance and prices.
    """
    start_dt, end_dt_excl, effective_to = _dt_bounds(from_date, to_date)
    start_date: date = start_dt.date()
    end_date_excl: date = end_dt_excl.date()

    # Filter set (optional single item)
    item_filter = []
    if item_id is not None:
        item_filter.append(Item.id == item_id)

    # 1) Determine which items have activity in the window (for pagination/count)
    items_in = (
        db.query(BatchStock.item_id)
        .join(Item, Item.id == BatchStock.item_id)
        .filter(
            BatchStock.tanggal_masuk >= start_date,
            BatchStock.tanggal_masuk < end_date_excl,
            *item_filter,
        )
    )
    items_out = (
        db.query(FifoLog.item_id)
        .join(Item, Item.id == FifoLog.item_id)
        .filter(
            FifoLog.invoice_date >= start_date,
            FifoLog.invoice_date < end_date_excl,
            *item_filter,
        )
    )
    active_item_ids = {rid[0] for rid in items_in.union(items_out).distinct().all()}

    total_count = len(active_item_ids)  # total items with activity
    if total_count == 0:
        title = f"Laporan Stock Adjustment {from_date:%d/%m/%Y} - {effective_to:%d/%m/%Y}"
        return StockAdjustmentReportResponse(
            title=title,
            date_from=from_date,
            date_to=effective_to,
            data=[],
            total=0,
        )

    # Pagination over item ids (deterministic order by item name)
    ordered_items = (
        db.query(Item.id, Item.name)
        .filter(Item.id.in_(active_item_ids))
        .order_by(Item.name.asc(), Item.id.asc())
        .offset(skip)
        .limit(limit)
        .all()
    )
    paged_item_ids = [row.id for row in ordered_items]

    # 2) Opening balance per item (before start_date)
    #    opening = sum(batches before) - sum(fifo before)
    pre_in_rows = (
        db.query(BatchStock.item_id, func.coalesce(func.sum(BatchStock.qty_masuk), 0))
        .filter(
            BatchStock.item_id.in_(paged_item_ids),
            BatchStock.tanggal_masuk < start_date,
        )
        .group_by(BatchStock.item_id)
        .all()
    )
    pre_out_rows = (
        db.query(FifoLog.item_id, func.coalesce(func.sum(FifoLog.qty_terpakai), 0))
        .filter(
            FifoLog.item_id.in_(paged_item_ids),
            FifoLog.invoice_date < start_date,
        )
        .group_by(FifoLog.item_id)
        .all()
    )
    pre_in_map = {r[0]: int(r[1]) for r in pre_in_rows}
    pre_out_map = {r[0]: int(r[1]) for r in pre_out_rows}

    opening_qty: Dict[int, int] = {
        iid: pre_in_map.get(iid, 0) - pre_out_map.get(iid, 0)
        for iid in paged_item_ids
    }

    # 3) Pull IN/OUT events inside window for paged items
    in_events = (
        db.query(
            BatchStock.item_id,
            Item.code.label("item_code"),
            Item.name.label("item_name"),
            BatchStock.id_batch,
            BatchStock.tanggal_masuk.label("event_date"),
            BatchStock.qty_masuk,
            BatchStock.harga_beli,
        )
        .join(Item, Item.id == BatchStock.item_id)
        .filter(
            BatchStock.item_id.in_(paged_item_ids),
            BatchStock.tanggal_masuk >= start_date,
            BatchStock.tanggal_masuk < end_date_excl,
        )
        .all()
    )

    out_events = (
        db.query(
            FifoLog.item_id,
            Item.code.label("item_code"),
            Item.name.label("item_name"),
            FifoLog.id_batch,
            FifoLog.invoice_date.label("event_date"),
            FifoLog.qty_terpakai,
            FifoLog.harga_modal,  # HPP per unit for that batch usage
            FifoLog.invoice_id,
        )
        .join(Item, Item.id == FifoLog.item_id)
        .filter(
            FifoLog.item_id.in_(paged_item_ids),
            FifoLog.invoice_date >= start_date,
            FifoLog.invoice_date < end_date_excl,
        )
        .all()
    )

    # 4) Build per-item timeline and running balance
    grouped_rows: Dict[str, List[StockAdjustmentReportRow]] = {}
    last_cost_per_item: Dict[int, Decimal] = {}

    # Seed last cost with last IN price before start (if any)
    seed_cost_rows = (
        db.query(BatchStock.item_id, BatchStock.harga_beli, BatchStock.tanggal_masuk)
        .filter(
            BatchStock.item_id.in_(paged_item_ids),
            BatchStock.tanggal_masuk < start_date,
        )
        .order_by(BatchStock.item_id.asc(), BatchStock.tanggal_masuk.desc(), BatchStock.id_batch.desc())
        .all()
    )
    for iid, harga_beli, _ in seed_cost_rows:
        if iid not in last_cost_per_item:
            last_cost_per_item[iid] = _D(harga_beli)

    # Compose events: IN and OUT
    per_item_events: Dict[int, List[dict]] = {iid: [] for iid in paged_item_ids}

    for ev in in_events:
        per_item_events[ev.item_id].append({
            "kind": "IN",
            "date": ev.event_date,
            "item_code": ev.item_code,
            "item_name": ev.item_name,
            "id_batch": ev.id_batch,
            "qty": int(ev.qty_masuk),
            "unit_cost": _D(ev.harga_beli),
            "ref": ev.id_batch,  # use batch as transaction no surrogate
            "no": f"BATCH-{ev.id_batch}",
        })

    for ev in out_events:
        per_item_events[ev.item_id].append({
            "kind": "OUT",
            "date": ev.event_date,
            "item_code": ev.item_code,
            "item_name": ev.item_name,
            "id_batch": ev.id_batch,
            "qty": int(ev.qty_terpakai),
            "unit_cost": _D(ev.harga_modal),  # HPP per unit actually used
            "ref": ev.invoice_id,
            "no": ev.invoice_id,
        })

    # Sort events per item (date, then IN before OUT to reflect stock arrival first)
    for iid in per_item_events:
        per_item_events[iid].sort(key=lambda x: (x["date"], 0 if x["kind"] == "IN" else 1, x.get("id_batch", 0)))

    # Build rows
    for iid in paged_item_ids:
        events = per_item_events[iid]
        if not events:
            # No rows in the window for this item (rare if active set correct)
            continue

        # Fetch item identity
        item_meta = next((e for e in events if e.get("item_name")), None)
        item_name = item_meta["item_name"] if item_meta else f"ITEM-{iid}"

        running = opening_qty.get(iid, 0)
        last_cost = last_cost_per_item.get(iid, Decimal("0"))

        for ev in events:
            if ev["kind"] == "IN":
                qty_in = ev["qty"]
                qty_out = 0
                harga_masuk = ev["unit_cost"]
                harga_keluar = Decimal("0")
                last_cost = harga_masuk  # update last known cost on purchase
                running += qty_in
                harga_beli = harga_masuk
                hpp = Decimal("0")
            else:
                qty_in = 0
                qty_out = ev["qty"]
                harga_masuk = Decimal("0")
                harga_keluar = ev["unit_cost"]  # actual HPP used for the OUT
                running -= qty_out
                # If we never saw an IN, fallback cost = this HPP
                if last_cost == 0:
                    last_cost = harga_keluar
                harga_beli = harga_keluar
                hpp = harga_keluar

            nilai_persediaan = _D(running) * _D(last_cost)

            row = StockAdjustmentReportRow(
                date=datetime.combine(ev["date"], time.min),
                no_transaksi=str(ev["no"]),
                batch=f"BATCH-{ev.get('id_batch')}" if ev.get("id_batch") else "N/A",
                item_code=ev["item_code"] or "N/A",
                item_name=ev["item_name"] or "N/A",
                qty_masuk=_D(qty_in),
                qty_keluar=_D(qty_out),
                qty_balance=_D(running),
                harga_masuk=_D(harga_masuk),
                harga_keluar=_D(harga_keluar),
                harga_beli=_D(harga_beli),
                nilai_persediaan=_D(nilai_persediaan),
                hpp=_D(hpp),
            )
            grouped_rows.setdefault(item_name, []).append(row)

    # Transform to response model
    items_payload: List[ItemStockAdjustmentReportRow] = [
        ItemStockAdjustmentReportRow(item_name=name, data=rows)
        for name, rows in grouped_rows.items()
    ]

    title = f"Laporan Stock Adjustment {from_date:%d/%m/%Y} - {effective_to:%d/%m/%Y}"
    return StockAdjustmentReportResponse(
        title=title,
        date_from=from_date,
        date_to=effective_to,
        data=items_payload,
        total=total_count,
    )

# ---------- XLSX download

@router.get(
    "/stock-adjustment/download",
    status_code=status.HTTP_200_OK,
    summary="Download Stock Adjustment Report as XLSX (from BatchStock + FifoLog)",
)
async def download_stock_adjustment_report(
    from_date: datetime = Query(..., description="Start datetime (inclusive)"),
    to_date: Optional[datetime] = Query(None, description="End datetime (inclusive)"),
    item_id: Optional[int] = Query(None, description="Filter by specific item"),
    db: Session = Depends(get_db),
):
    start_dt, end_dt_excl, effective_to = _dt_bounds(from_date, to_date)
    start_date: date = start_dt.date()
    end_date_excl: date = end_dt_excl.date()

    item_filter = []
    if item_id is not None:
        item_filter.append(Item.id == item_id)

    # Gather ALL items with activity (no pagination in export)
    items_in = (
        db.query(BatchStock.item_id)
        .join(Item, Item.id == BatchStock.item_id)
        .filter(
            BatchStock.tanggal_masuk >= start_date,
            BatchStock.tanggal_masuk < end_date_excl,
            *item_filter,
        )
    )
    items_out = (
        db.query(FifoLog.item_id)
        .join(Item, Item.id == FifoLog.item_id)
        .filter(
            FifoLog.invoice_date >= start_date,
            FifoLog.invoice_date < end_date_excl,
            *item_filter,
        )
    )
    item_ids = {rid[0] for rid in items_in.union(items_out).distinct().all()}
    if not item_ids:
        # produce empty sheet with headers
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Adjustment"
        ws.append(["No data for the selected period."])
        out = io.BytesIO(); wb.save(out); out.seek(0)
        filename = f"laporan_stock_adjustment_{from_date:%Y%m%d}_{effective_to:%Y%m%d}.xlsx"
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # Opening balances
    pre_in_rows = (
        db.query(BatchStock.item_id, func.coalesce(func.sum(BatchStock.qty_masuk), 0))
        .filter(BatchStock.item_id.in_(item_ids), BatchStock.tanggal_masuk < start_date)
        .group_by(BatchStock.item_id)
        .all()
    )
    pre_out_rows = (
        db.query(FifoLog.item_id, func.coalesce(func.sum(FifoLog.qty_terpakai), 0))
        .filter(FifoLog.item_id.in_(item_ids), FifoLog.invoice_date < start_date)
        .group_by(FifoLog.item_id)
        .all()
    )
    pre_in_map = {r[0]: int(r[1]) for r in pre_in_rows}
    pre_out_map = {r[0]: int(r[1]) for r in pre_out_rows}
    opening_qty = {iid: pre_in_map.get(iid, 0) - pre_out_map.get(iid, 0) for iid in item_ids}

    # Seed last cost with most recent prior purchase per item
    seed_cost_rows = (
        db.query(BatchStock.item_id, BatchStock.harga_beli, BatchStock.tanggal_masuk)
        .filter(BatchStock.item_id.in_(item_ids), BatchStock.tanggal_masuk < start_date)
        .order_by(BatchStock.item_id.asc(), BatchStock.tanggal_masuk.desc(), BatchStock.id_batch.desc())
        .all()
    )
    last_cost = {}
    for iid, harga_beli, _ in seed_cost_rows:
        if iid not in last_cost:
            last_cost[iid] = _D(harga_beli)

    # Pull events
    in_events = (
        db.query(
            BatchStock.item_id, Item.code, Item.name,
            BatchStock.id_batch, BatchStock.tanggal_masuk.label("d"),
            BatchStock.qty_masuk, BatchStock.harga_beli
        )
        .join(Item, Item.id == BatchStock.item_id)
        .filter(
            BatchStock.item_id.in_(item_ids),
            BatchStock.tanggal_masuk >= start_date,
            BatchStock.tanggal_masuk < end_date_excl,
        ).all()
    )
    out_events = (
        db.query(
            FifoLog.item_id, Item.code, Item.name,
            FifoLog.id_batch, FifoLog.invoice_date.label("d"),
            FifoLog.qty_terpakai, FifoLog.harga_modal, FifoLog.invoice_id
        )
        .join(Item, Item.id == FifoLog.item_id)
        .filter(
            FifoLog.item_id.in_(item_ids),
            FifoLog.invoice_date >= start_date,
            FifoLog.invoice_date < end_date_excl,
        ).all()
    )

    per_item = {iid: [] for iid in item_ids}
    for ev in in_events:
        per_item[ev.item_id].append(("IN", ev.d, ev.code, ev.name, ev.id_batch, int(ev.qty_masuk), _D(ev.harga_beli), f"BATCH-{ev.id_batch}"))
    for ev in out_events:
        per_item[ev.item_id].append(("OUT", ev.d, ev.code, ev.name, ev.id_batch, int(ev.qty_terpakai), _D(ev.harga_modal), ev.invoice_id))

    for iid in per_item:
        per_item[iid].sort(key=lambda x: (x[1], 0 if x[0] == "IN" else 1, x[4] or 0))

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Adjustment"
    headers = [
        "Date","No Transaksi","Batch","Item Code","Item Name",
        "Qty Masuk","Qty Keluar","Qty Balance",
        "Harga Masuk","Harga Keluar","Harga Beli","Nilai Persediaan","HPP (OUT)"
    ]
    ws.append(headers)

    # For deterministic order
    item_order = (
        db.query(Item.id, Item.name).filter(Item.id.in_(item_ids)).order_by(Item.name.asc(), Item.id.asc()).all()
    )

    for iid, iname in item_order:
        running = opening_qty.get(iid, 0)
        cost = last_cost.get(iid, Decimal("0"))
        events = per_item[iid]

        for kind, d, code, name, id_batch, qty, unit_cost, ref in events:
            if kind == "IN":
                qty_in, qty_out = qty, 0
                harga_masuk, harga_keluar = unit_cost, Decimal("0")
                cost = unit_cost
                running += qty_in
                hpp_out = Decimal("0")
                harga_beli = unit_cost
            else:
                qty_in, qty_out = 0, qty
                harga_masuk, harga_keluar = Decimal("0"), unit_cost
                running -= qty_out
                if cost == 0:
                    cost = unit_cost
                hpp_out = unit_cost
                harga_beli = unit_cost

            nilai_persediaan = _D(running) * _D(cost)
            ws.append([
                d.strftime("%d/%m/%Y"),
                str(ref),
                f"BATCH-{id_batch}" if id_batch else "N/A",
                code or "N/A",
                name or "N/A",
                float(_D(qty_in)),
                float(_D(qty_out)),
                float(_D(running)),
                float(_D(harga_masuk)),
                float(_D(harga_keluar)),
                float(_D(harga_beli)),
                float(_D(nilai_persediaan)),
                float(_D(hpp_out)),
            ])

    # Autofit-ish
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"laporan_stock_adjustment_{from_date:%Y%m%d}_{effective_to:%Y%m%d}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )

from datetime import datetime
import pytz

@router.post("/migrate-batch-stocks")
async def migrate_batch_stocks_for_all_items(
    db: Session = Depends(get_db),
):
    """
    ONE-TIME MIGRATION: Create initial BatchStock records for all existing items.
    
    This endpoint:
    - Processes ALL items (including deleted, zero stock, and service items)
    - Creates one BatchStock per item using current stock as qty_masuk
    - Uses current price (not modal_price) as harga_beli
    - Sets source_type = ITEM, source_id = item.id
    - Uses current JKT time as tanggal_masuk
    - Creates audit log for each item
    
    Returns summary of migration results.
    """
    
    try:
        # Get JKT timezone
        jkt_tz = pytz.timezone('Asia/Jakarta')
        current_jkt_date = datetime.now(jkt_tz).date()
        
        audit_service = AuditService(db)
        
        # Get ALL items (no filters)
        all_items = db.query(Item).all()
        
        if not all_items:
            return {
                "success": False,
                "message": "No items found in database",
                "total_items": 0,
                "batches_created": 0,
                "errors": []
            }
        
        total_items = len(all_items)
        batches_created = 0
        errors = []
        
        print(f"\n{'='*60}")
        print(f"Starting Batch Stock Migration")
        print(f"Total items to process: {total_items}")
        print(f"Current JKT Date: {current_jkt_date}")
        print(f"{'='*60}\n")
        
        for idx, item in enumerate(all_items, 1):
            try:
                # Use current price (not modal_price) as harga_beli
                harga_beli = Decimal(str(item.price)) if item.price else Decimal('0')
                qty_masuk = item.total_item if item.total_item else 0
                nilai_total = Decimal(qty_masuk) * harga_beli
                
                # Create batch with source_type = ITEM (from SourceTypeEnum)
                batch = BatchStock(
                    source_id=str(item.id),  # Use item.id as source_id
                    source_type=SourceTypeEnum.ITEM,  # Use ITEM enum value
                    item_id=item.id,
                    warehouse_id=None,  # No warehouse
                    tanggal_masuk=current_jkt_date,
                    qty_masuk=qty_masuk,
                    qty_keluar=0,
                    sisa_qty=qty_masuk,
                    harga_beli=harga_beli,
                    nilai_total=nilai_total,
                    is_open=True
                )
                
                db.add(batch)
                batches_created += 1
                
                # Create audit log
                audit_service.default_log(
                    entity_id=item.id,
                    entity_type=AuditEntityEnum.ITEM,
                    description=f"Initial batch stock created for item {item.name} (qty: {qty_masuk}, price: {harga_beli})",
                    user_name="ADMIN"
                )
                
                # Print progress every 10 items
                if idx % 10 == 0 or idx == total_items:
                    print(f"Progress: {idx}/{total_items} items processed...")
                
            except Exception as e:
                error_msg = f"Failed to create batch for item {item.id} ({item.name}): {str(e)}"
                errors.append({
                    "item_id": item.id,
                    "item_name": item.name,
                    "sku": item.sku,
                    "error": str(e)
                })
                print(f"ERROR: {error_msg}")
                continue
        
        # Commit all changes
        db.commit()
        
        print(f"\n{'='*60}")
        print(f"Migration Complete!")
        print(f"Total items: {total_items}")
        print(f"Batches created: {batches_created}")
        print(f"Errors: {len(errors)}")
        print(f"{'='*60}\n")
        
        return {
            "success": True,
            "message": "Batch stock migration completed successfully",
            "total_items": total_items,
            "batches_created": batches_created,
            "failed_items": len(errors),
            "errors": errors if errors else None,
            "migration_date": current_jkt_date.isoformat()
        }
        
    except Exception as e:
        db.rollback()
        print(f"\nFATAL ERROR: {str(e)}\n")
        raise HTTPException(
            status_code=500,
            detail=f"Migration failed: {str(e)}"
        )