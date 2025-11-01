# Standard Library Imports
import io
import os
import shutil
import uuid
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any, Dict, List, Literal, Optional

# ---

# Third-Party Imports
import pandas as pd
from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    Query,
    Request,
    UploadFile,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import BaseModel  # Added BaseModel from the duplicates
from sqlalchemy import inspect, or_, text
from sqlalchemy.orm import Session, joinedload
from starlette.exceptions import HTTPException
from starlette.responses import StreamingResponse

# ---

# Local/Application-Specific Imports
from database import get_db
from models.InventoryLedger import SourceTypeEnum
from models.Satuan import  Satuan
from models.Category import Category
from models.AllAttachment import AllAttachment, ParentType
from models.AuditTrail import AuditEntityEnum

from models.Item import Item
from models.Vendor import Vendor
from routes.category_routes import _build_categories_lookup
from routes.satuan_routes import _build_satuans_lookup
from routes.vendor_routes import _build_vendors_lookup
from schemas.CategorySchemas import CategoryOut
from schemas.ItemSchema import ItemResponse, ItemTypeEnum
from schemas.PaginatedResponseSchemas import PaginatedResponse
from schemas.SatuanSchemas import SatuanOut
from schemas.VendorSchemas import VendorOut
from services.audit_services import AuditService
from services.fifo_services import FifoService

from utils import (
    generate_unique_record_code,
    get_current_user_name,
    soft_delete_record,
)
router = APIRouter()

NEXT_PUBLIC_UPLOAD_DIR = os.getenv("UPLOAD_DIR", default="uploads/items")
os.makedirs(NEXT_PUBLIC_UPLOAD_DIR, exist_ok=True)


def get_item_prefix(item_type: ItemTypeEnum) -> str:
    if item_type == ItemTypeEnum.HIGH_QUALITY:
        return "FG"
    elif item_type == ItemTypeEnum.RAW_MATERIAL:
        return "RAW"
    elif item_type == ItemTypeEnum.SERVICE:
        return "SERVICE"
    else:
        raise ValueError(f"Unsupported item type: {item_type}")




class ImportResult(BaseModel):
    total_processed: int
    successful_imports: int
    failed_imports: int
    errors: List[Dict[str, Any]]
    warnings: List[Dict[str, Any]]

class ImportOptions(BaseModel):
    skip_on_error: bool = True
    update_existing: bool = False
    default_item_type: ItemTypeEnum = ItemTypeEnum.HIGH_QUALITY




@router.get("/export-excel")
async def export_items_to_excel(
        db: Session = Depends(get_db),
        item_type: Optional[ItemTypeEnum] = Query(None, description="Filter by item type"),
        is_active: Optional[bool] = Query(None, description="Filter by active status"),
        include_inactive: bool = Query(False, description="Include inactive items"),
):
    """
    Export all items to Excel file using the same template format as import.
    
    Columns exported:
    - Type
    - Nama Item
    - SKU
    - Brand
    - Jenis Barang
    - Jumlah Unit
    - Harga Modal
    - Harga Jual
    - Satuan Unit
    """
    
    try:
        # Build query with filters
        query = db.query(Item).filter(Item.deleted_at.is_(None))
        
        if item_type:
            query = query.filter(Item.type == item_type)
        
        if is_active is not None:
            query = query.filter(Item.is_active == is_active)
        elif not include_inactive:
            query = query.filter(Item.is_active == True)
        
        # Get all items
        items = query.order_by(Item.code).all()
        
        if not items:
            raise HTTPException(
                status_code=404,
                detail="No items found to export"
            )
        
        # Build lookup dictionaries for efficient data retrieval
        category_ids = set()
        satuan_ids = set()
        
        for item in items:
            if item.category_one:
                category_ids.add(item.category_one)
            if item.category_two:
                category_ids.add(item.category_two)
            if item.satuan_id:
                satuan_ids.add(item.satuan_id)
        
        # Fetch categories and satuans
        categories = {}
        if category_ids:
            cats = db.query(Category).filter(Category.id.in_(category_ids)).all()
            categories = {cat.id: cat.name for cat in cats}
        
        satuans = {}
        if satuan_ids:
            sats = db.query(Satuan).filter(Satuan.id.in_(satuan_ids)).all()
            satuans = {sat.id: sat.symbol for sat in sats}
        
        # Map ItemTypeEnum to readable format
        type_mapping = {
            ItemTypeEnum.HIGH_QUALITY: "High Quality",
            ItemTypeEnum.RAW_MATERIAL: "Raw Material",
            ItemTypeEnum.SERVICE: "Service"
        }
        
        # Prepare data for DataFrame
        data = []
        for item in items:
            data.append({
                'Type': type_mapping.get(item.type, str(item.type)),
                'Nama Item': item.name,
                'SKU': item.sku,
                'Brand': categories.get(item.category_one, ''),
                'Jenis Barang': categories.get(item.category_two, ''),
                'Jumlah Unit': item.total_item if item.total_item else 0,
                'Harga Modal': float(item.modal_price) if item.modal_price else 0,
                'Harga Jual': float(item.price) if item.price else 0,
                'Satuan Unit': satuans.get(item.satuan_id, '')
            })
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create Excel file with formatting
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write the header row
            df.to_excel(writer, index=False, sheet_name='Items', startrow=0)
            
            # Get the worksheet to add notes row
            worksheet = writer.sheets['Items']
            
            # Insert a row for notes after the header
            worksheet.insert_rows(2)
            notes = [
                'e.g., High Quality, Raw Material, Service',
                'Required: Item name',
                'Required: Unique identifier',
                'Optional: Brand category name',
                'Optional: Item category name',
                'Optional: Current stock quantity',
                'Optional: Cost price',
                'Required: Selling price',
                'Required: Unit symbol (e.g., pcs, kg, m)'
            ]
            
            for col_idx, note in enumerate(notes, start=1):
                cell = worksheet.cell(row=2, column=col_idx)
                cell.value = note
                cell.font = cell.font.copy(italic=True, size=9)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"items_export_{timestamp}.xlsx"
        
        # Return as streaming response
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error exporting items: {str(e)}"
        )




@router.get("/{item_id}", response_model=ItemResponse)
def get_item_by_id(
        request: Request,
        item_id: int,
        db: Session = Depends(get_db),
):
    db_item = db.query(Item).options(
        joinedload(Item.category_one_rel),
        joinedload(Item.category_two_rel),
        joinedload(Item.satuan_rel),
        joinedload(Item.attachments),
    ).filter(Item.id == item_id, Item.is_deleted == False).first()

    if not db_item:
        raise HTTPException(status_code=404, detail="Item tidak ditemukan")

    return construct_item_response(db_item, request)

@router.post("", response_model=ItemResponse)
async def create_item(
        images: List[UploadFile] = File(default=[]),
        type: ItemTypeEnum = Form(...),
        name: str = Form(...),
        sku: str = Form(...),
        total_item: int = Form(0),
        min_item: int = Form(0),
        price: float = Form(...),
        modal_price: float = Form(...),
        is_active: bool = Form(True),
        category_one: Optional[int] = Form(None),
        category_two: Optional[int] = Form(None),
        vendor_id: Optional[str] = Form(None),
        satuan_id: int = Form(...),
        db: Session = Depends(get_db),
        user_name: str = Depends(get_current_user_name),
):
    if len(images) > 3:
        raise HTTPException(status_code=400, detail="Maximum 3 images allowed")

    pattern = get_item_prefix(type)

    audit_service = AuditService(db)

    # SKU validation
    existing_item = db.query(Item).filter(Item.sku == sku).first()
    if existing_item:
        if existing_item.is_deleted:
            existing_item.type = type
            existing_item.code = generate_unique_record_code(db, Item, pattern)
            existing_item.name = name
            existing_item.total_item = total_item
            existing_item.min_item = min_item
            existing_item.price = price
            existing_item.modal_price = modal_price
            existing_item.is_active = is_active
            existing_item.category_one = category_one
            existing_item.category_two = category_two
            existing_item.satuan_id = satuan_id
            existing_item.vendor_id = vendor_id
            existing_item.is_deleted = False
            existing_item.deleted_at = None
            db.commit()
            db.refresh(existing_item)
            return existing_item
        else:
            raise HTTPException(status_code=400, detail="SKU sudah ada")

    try:
        db_item = Item(
            type=type,
            name=name,
            code=generate_unique_record_code(db, Item, pattern),
            sku=sku,
            total_item=total_item,
            min_item=min_item,
            price=price,
            modal_price=modal_price,
            is_active=is_active,
            category_one=category_one,
            category_two=category_two,
            vendor_id=vendor_id,
            satuan_id=satuan_id,
        )

        db.add(db_item)
        db.commit()
        db.refresh(db_item)

        # Handle images
        for image in images:
            if image.filename:
                ext = os.path.splitext(image.filename)[1]
                unique_filename = f"{uuid.uuid4()}{ext}"
                save_path = os.path.join(NEXT_PUBLIC_UPLOAD_DIR, unique_filename)

                with open(save_path, "wb") as buffer:
                    shutil.copyfileobj(image.file, buffer)

                attachment = AllAttachment(
                    parent_type=ParentType.ITEMS,
                    item_id=db_item.id,
                    filename=image.filename,
                    file_path=save_path,
                    file_size=os.path.getsize(save_path),
                    mime_type=image.content_type,
                    created_at=datetime.now(),
                )
                db.add(attachment)

        db.commit()
        db.refresh(db_item)

        audit_service.default_log(
            entity_id=db_item.id,
            entity_type=AuditEntityEnum.ITEM,
            description=f"Item {db_item.name} telah dibuat",
            user_name=user_name,
        )

        return db_item

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"{str(e)}")
@router.get("", response_model=PaginatedResponse[ItemResponse])
def get_items(
        request: Request,
        db: Session = Depends(get_db),
        page: int = 1,
        rowsPerPage: int = 5,
        search_key: Optional[str] = None,

        vendor: Optional[str] = None,
        item_type: Optional[ItemTypeEnum] = None,
        
        contains_deleted: Optional[bool] = False,
        is_active: Optional[bool] = None,
        sortBy: Optional[Literal["name", "price", "sku", "created_at"]] = None,
        sortOrder: Optional[Literal["asc", "desc"]] = "asc",
        to_date : Optional[date] = Query(None, description="Filter by date"),
        from_date : Optional[date] = Query(None, description="Filter by date")

):
    query = db.query(Item).options(
        joinedload(Item.category_one_rel),
        joinedload(Item.category_two_rel),
        joinedload(Item.vendor_rel),
        joinedload(Item.satuan_rel),
        joinedload(Item.attachments),
    )
    # Remove the .order_by(Item.created_at.desc()) from here
    
    if contains_deleted is False:
        query = query.filter(Item.is_deleted == False)

    if search_key:
        query = query.filter(
            or_(
                Item.name.ilike(f"%{search_key}%"),
                Item.sku.ilike(f"%{search_key}%"),
            )
        )
    if vendor and vendor != "all":
        query = query.filter(Item.vendor_id == vendor)


    if item_type:
        query = query.filter(Item.type == item_type)

    if is_active is not None:
        query = query.filter(Item.is_active == is_active)

    if from_date and to_date:
        query = query.filter(
            Item.created_at.between(
                datetime.combine(from_date, time.min),
                datetime.combine(to_date, time.max),
            )
        )
    elif from_date:
        query = query.filter(Item.created_at >= datetime.combine(from_date, time.min))
    elif to_date:
        query = query.filter(Item.created_at <= datetime.combine(to_date, time.max))

    # Apply sorting - either custom or default
    if sortBy:
        sort_column = getattr(Item, sortBy)
        query = query.order_by(sort_column.desc() if sortOrder == "desc" else sort_column.asc())
    else:
        # Default sorting when no sortBy is specified
        query = query.order_by(Item.created_at.desc())

    total_count = query.count()
    paginated_data = query.offset((page - 1) * rowsPerPage).limit(rowsPerPage).all()

    items_out = [construct_item_response(item, request) for item in paginated_data]

    return {"data": items_out, "total": total_count}

def _update_existing_item(db: Session, item_data: Dict[str, Any], existing_item_id: int, audit_service: AuditService, user_name: str):
    """Update an existing item without changing its code."""
   

    item = db.query(Item).filter(Item.id == existing_item_id).first()
    if not item:
        raise ValueError(f"Item with ID {existing_item_id} not found")

    old_total_item = item.total_item

    # Update ONLY the allowed fields - NEVER touch 'code', 'id', 'created_at'
    updateable_fields = ['name', 'type', 'total_item', 'price', 'vendor_id',
                         'category_one', 'category_two', 'satuan_id', 'is_active']

    for field in updateable_fields:
        if field in item_data:
            setattr(item, field, item_data[field])

    new_total_item = item_data.get('total_item', 0)
    if new_total_item != old_total_item:
        difference = new_total_item - old_total_item

        unique_source_id = f"IMPORT-{item.sku}"
        
        if difference > 0:
            # Stock increase - create FIFO batch
            FifoService.create_batch_from_purchase(
                source_id=str(item.id),
                source_type=SourceTypeEnum.ITEM,
                db=db,
                item_id=item.id,
                warehouse_id=None,  # No warehouse for import
                tanggal_masuk=date.today(),
                qty_masuk=difference,
                harga_beli=item_data.get('price', Decimal('0'))
            )
            
        elif difference < 0:
            # Stock decrease - consume from FIFO batches
            try:
                FifoService.process_sale_fifo(
                    db=db,
                    invoice_id=unique_source_id,
                    invoice_date=date.today(),
                    item_id=item.id,
                    qty_terjual=abs(difference),
                    harga_jual_per_unit=Decimal("0"),  # No sale price for import adjustments
                    warehouse_id=None
                )
            except ValueError as e:
                raise ValueError(f"Cannot reduce stock for {item.name}: {str(e)}")

    audit_service.default_log(
        entity_id=item.id,
        entity_type=AuditEntityEnum.ITEM,
        description=f"Data item {item.name} telah diupdate via import",
        user_name=user_name,
    )

    # Flush changes to the database
    db.flush()

    return item

def _create_new_item(db: Session, item_data: Dict[str, Any], audit_service: AuditService, user_name: str):
    """Create a new item and post initial inventory using FIFO."""
    
    new_item = Item(**item_data)
    db.add(new_item)
    db.flush()

    # Create initial FIFO batch if there's stock
    if item_data.get('total_item', 0) > 0:
        unit_price = item_data.get('price', Decimal('0'))
        qty = item_data['total_item']
        
        FifoService.create_batch_from_purchase(
            source_id=str(new_item.id),
            source_type=SourceTypeEnum.ITEM,
            db=db,
            item_id=new_item.id,
            warehouse_id=None,  # No warehouse for initial import
            tanggal_masuk=date.today(),
            qty_masuk=qty,
            harga_beli=unit_price
        )

    audit_service.default_log(
        entity_id=new_item.id,
        entity_type=AuditEntityEnum.ITEM,
        description=f"Data item {new_item.name} telah dibuat via import",
        user_name=user_name,
    )
    return new_item

@router.get("/template/download")
async def download_item_template(format: str = "xlsx"):
    """
    Download item import template in Excel format.

    Features:
    - Proper number formatting for prices and quantities
    - Data validation dropdowns for Type column
    - Helper notes for field usage
    - UTF-8 encoding support
    """

    if format.lower() != "xlsx":
        raise HTTPException(
            status_code=400,
            detail="Only xlsx format is supported. Use format=xlsx"
        )

    try:
        # Create workbook with two sheets
        wb = Workbook()

        # Sheet 1: Main template
        ws_template = wb.active
        ws_template.title = "Item_Template"

        # Sheet 2: Dropdown reference
        ws_dropdown = wb.create_sheet("Dropdown_Reference")

        # === SHEET 1: Item_Template ===

        # Define headers - ADDED Stock Minimum
        headers = [
            "Type",
            "Nama Item",
            "SKU",
            "Brand",
            "Jenis Barang",
            "Jumlah Unit",
            "Stock Minimum",  # NEW COLUMN
            "Harga Modal",
            "Harga Jual",
            "Satuan Unit",
            "Nama Vendor"
        ]

        # Style definitions
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        note_font = Font(name='Calibri', size=9, italic=True, color='7F7F7F')
        note_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        border_thin = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # Write headers (Row 1)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_template.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_thin

        # Write helper notes (Row 2) - ADDED Stock Minimum note
        notes = [
            "High Quality / Raw Material / Service",
            "Nama produk/barang (wajib diisi)",
            "Kode unik produk (wajib diisi, tidak boleh duplikat)",
            "Merek produk (opsional, harus sudah terdaftar)",
            "Kategori barang (opsional, harus sudah terdaftar)",
            "Jumlah stok awal (angka, default: 0)",
            "Batas minimal stok untuk notifikasi (angka, default: 0)",  # NEW NOTE
            "Harga pokok/modal (angka, default: 0)",
            "Harga jual (angka, wajib diisi)",
            "Satuan unit (wajib diisi, harus sudah terdaftar, contoh: pcs, kg, box)",
            "Nama Vendor (opsional, harus sudah terdaftar, contoh: Vendor A)"
        ]

        for col_idx, note in enumerate(notes, start=1):
            cell = ws_template.cell(row=2, column=col_idx, value=note)
            cell.font = note_font
            cell.fill = note_fill
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = border_thin

        # Set column widths - ADDED Stock Minimum width
        column_widths = {
            'A': 18,  # Type
            'B': 30,  # Nama Item
            'C': 15,  # SKU
            'D': 15,  # Brand
            'E': 18,  # Jenis Barang
            'F': 14,  # Jumlah Unit
            'G': 15,  # Stock Minimum (NEW)
            'H': 15,  # Harga Modal
            'I': 15,  # Harga Jual
            'J': 15,  # Satuan Unit
            'K': 18   # Nama Vendor
        }

        for col, width in column_widths.items():
            ws_template.column_dimensions[col].width = width

        # Set row heights
        ws_template.row_dimensions[1].height = 30
        ws_template.row_dimensions[2].height = 45

        # Add sample data (Row 3) with proper formatting - ADDED Stock Minimum sample
        sample_data = [
            "High Quality",
            "Contoh Produk A",
            "SKU-001",
            "Brand A",
            "Elektronik",
            100,      # Jumlah Unit
            10,       # Stock Minimum (NEW)
            50000,    # Harga Modal
            75000,    # Harga Jual
            "pcs",
            "Vendor A"
        ]

        for col_idx, value in enumerate(sample_data, start=1):
            cell = ws_template.cell(row=3, column=col_idx, value=value)
            cell.border = border_thin

            # Apply number format to numeric columns
            if col_idx in [6, 7, 8, 9]:  # Jumlah Unit, Stock Minimum, Harga Modal, Harga Jual
                cell.number_format = '#,##0'

        # Format numeric columns for the entire range (rows 3-1000)
        for row_idx in range(3, 1001):
            # Jumlah Unit (F)
            ws_template.cell(row=row_idx, column=6).number_format = '#,##0'
            # Stock Minimum (G) - NEW
            ws_template.cell(row=row_idx, column=7).number_format = '#,##0'
            # Harga Modal (H)
            ws_template.cell(row=row_idx, column=8).number_format = '#,##0'
            # Harga Jual (I)
            ws_template.cell(row=row_idx, column=9).number_format = '#,##0'

        # === SHEET 2: Dropdown_Reference ===

        # Add dropdown values
        dropdown_values = ["High Quality", "Raw Material", "Service"]

        ws_dropdown.cell(row=1, column=1, value="Item Types")
        ws_dropdown.cell(row=1, column=1).font = Font(bold=True)

        for idx, value in enumerate(dropdown_values, start=2):
            ws_dropdown.cell(row=idx, column=1, value=value)

        ws_dropdown.column_dimensions['A'].width = 20

        # === DATA VALIDATION for Type column ===

        # Create data validation for Type column
        dv = DataValidation(
            type="list",
            formula1="Dropdown_Reference!$A$2:$A$4",
            allow_blank=True
        )
        dv.error = "Pilih salah satu: High Quality, Raw Material, atau Service"
        dv.errorTitle = "Input Tidak Valid"
        dv.prompt = "Pilih tipe item dari dropdown"
        dv.promptTitle = "Tipe Item"

        # Apply validation to Type column (A3:A1000)
        ws_template.add_data_validation(dv)
        dv.add("A3:A1000")

        # Freeze panes (freeze first 2 rows)
        ws_template.freeze_panes = "A3"

        # Hide Dropdown_Reference sheet
        ws_dropdown.sheet_state = 'hidden'

        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Return as streaming response
        headers = {
            'Content-Disposition': 'attachment; filename="Template_Import_Item.xlsx"',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }

        return StreamingResponse(
            excel_file,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error generating Excel template: {str(e)}"
        )

        
@router.get("/template/download")
async def download_item_template(format: str = "xlsx"):
    """
    Download item import template in Excel format.

    Features:
    - Proper number formatting for prices and quantities
    - Data validation dropdowns for Type column
    - Helper notes for field usage
    - UTF-8 encoding support
    """

    if format.lower() != "xlsx":
        raise HTTPException(
            status_code=400,
            detail="Only xlsx format is supported. Use format=xlsx"
        )

    try:
        # Create workbook with two sheets
        wb = Workbook()

        # Sheet 1: Main template
        ws_template = wb.active
        ws_template.title = "Item_Template"

        # Sheet 2: Dropdown reference
        ws_dropdown = wb.create_sheet("Dropdown_Reference")

        # === SHEET 1: Item_Template ===

        # Define headers - ADDED Stock Minimum
        headers = [
            "Type",
            "Nama Item",
            "SKU",
            "Brand",
            "Jenis Barang",
            "Jumlah Unit",
            "Stock Minimum",  # NEW COLUMN
            "Harga Modal",
            "Harga Jual",
            "Satuan Unit",
            "Nama Vendor"
        ]

        # Style definitions
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        note_font = Font(name='Calibri', size=9, italic=True, color='7F7F7F')
        note_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        border_thin = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # Write headers (Row 1)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_template.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_thin

        # Write helper notes (Row 2) - ADDED Stock Minimum note
        notes = [
            "High Quality / Raw Material / Service",
            "Nama produk/barang (wajib diisi)",
            "Kode unik produk (wajib diisi, tidak boleh duplikat)",
            "Merek produk (opsional, harus sudah terdaftar)",
            "Kategori barang (opsional, harus sudah terdaftar)",
            "Jumlah stok awal (angka, default: 0)",
            "Batas minimal stok untuk notifikasi (angka, default: 0)",  # NEW NOTE
            "Harga pokok/modal (angka, default: 0)",
            "Harga jual (angka, wajib diisi)",
            "Satuan unit (wajib diisi, harus sudah terdaftar, contoh: pcs, kg, box)",
            "Nama Vendor (opsional, harus sudah terdaftar, contoh: Vendor A)"
        ]

        for col_idx, note in enumerate(notes, start=1):
            cell = ws_template.cell(row=2, column=col_idx, value=note)
            cell.font = note_font
            cell.fill = note_fill
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = border_thin

        # Set column widths - ADDED Stock Minimum width
        column_widths = {
            'A': 18,  # Type
            'B': 30,  # Nama Item
            'C': 15,  # SKU
            'D': 15,  # Brand
            'E': 18,  # Jenis Barang
            'F': 14,  # Jumlah Unit
            'G': 15,  # Stock Minimum (NEW)
            'H': 15,  # Harga Modal
            'I': 15,  # Harga Jual
            'J': 15,  # Satuan Unit
            'K': 18   # Nama Vendor
        }

        for col, width in column_widths.items():
            ws_template.column_dimensions[col].width = width

        # Set row heights
        ws_template.row_dimensions[1].height = 30
        ws_template.row_dimensions[2].height = 45

        # Add sample data (Row 3) with proper formatting - ADDED Stock Minimum sample
        sample_data = [
            "High Quality",
            "Contoh Produk A",
            "SKU-001",
            "Brand A",
            "Elektronik",
            100,      # Jumlah Unit
            10,       # Stock Minimum (NEW)
            50000,    # Harga Modal
            75000,    # Harga Jual
            "pcs",
            "Vendor A"
        ]

        for col_idx, value in enumerate(sample_data, start=1):
            cell = ws_template.cell(row=3, column=col_idx, value=value)
            cell.border = border_thin

            # Apply number format to numeric columns
            if col_idx in [6, 7, 8, 9]:  # Jumlah Unit, Stock Minimum, Harga Modal, Harga Jual
                cell.number_format = '#,##0'

        # Format numeric columns for the entire range (rows 3-1000)
        for row_idx in range(3, 1001):
            # Jumlah Unit (F)
            ws_template.cell(row=row_idx, column=6).number_format = '#,##0'
            # Stock Minimum (G) - NEW
            ws_template.cell(row=row_idx, column=7).number_format = '#,##0'
            # Harga Modal (H)
            ws_template.cell(row=row_idx, column=8).number_format = '#,##0'
            # Harga Jual (I)
            ws_template.cell(row=row_idx, column=9).number_format = '#,##0'

        # === SHEET 2: Dropdown_Reference ===

        # Add dropdown values
        dropdown_values = ["High Quality", "Raw Material", "Service"]

        ws_dropdown.cell(row=1, column=1, value="Item Types")
        ws_dropdown.cell(row=1, column=1).font = Font(bold=True)

        for idx, value in enumerate(dropdown_values, start=2):
            ws_dropdown.cell(row=idx, column=1, value=value)

        ws_dropdown.column_dimensions['A'].width = 20

        # === DATA VALIDATION for Type column ===

        # Create data validation for Type column
        dv = DataValidation(
            type="list",
            formula1="Dropdown_Reference!$A$2:$A$4",
            allow_blank=True
        )
        dv.error = "Pilih salah satu: High Quality, Raw Material, atau Service"
        dv.errorTitle = "Input Tidak Valid"
        dv.prompt = "Pilih tipe item dari dropdown"
        dv.promptTitle = "Tipe Item"

        # Apply validation to Type column (A3:A1000)
        ws_template.add_data_validation(dv)
        dv.add("A3:A1000")

        # Freeze panes (freeze first 2 rows)
        ws_template.freeze_panes = "A3"

        # Hide Dropdown_Reference sheet
        ws_dropdown.sheet_state = 'hidden'

        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Return as streaming response
        headers = {
            'Content-Disposition': 'attachment; filename="Template_Import_Item.xlsx"',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }

        return StreamingResponse(
            excel_file,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error generating Excel template: {str(e)}"
        )


@router.post("/import-excel", response_model=ImportResult)
async def import_items_from_excel(
        db: Session = Depends(get_db),
        file: UploadFile = File(...),
        skip_on_error: bool = Query(True, description="Skip rows with errors instead of failing completely"),
        update_existing: bool = Query(False, description="Update existing items if SKU already exists"),
        default_item_type: ItemTypeEnum = Query(ItemTypeEnum.HIGH_QUALITY, description="Default item type if not specified"),
        user_name: str = Depends(get_current_user_name)
):
    """
    Import items from Excel/CSV file using the template format.

    Expected columns:
    - Type (optional, uses default if not provided)
    - Nama Item (required)
    - SKU (required, unique)
    - Brand (optional, by name)
    - Jenis Barang (optional, by name)
    - Jumlah Unit (optional, defaults to 0)
    - Stock Minimum (optional, defaults to 0) - NEW
    - Harga Modal (optional, defaults to 0)
    - Harga Jual (required)
    - Satuan Unit (required, by name)
    - Nama Vendor (optional, by name)

    Note: Item Code will be auto-generated based on item type
    """

    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(
            status_code=400,
            detail="File must be Excel (.xlsx, .xls) or CSV (.csv)"
        )

    try:
        # Read file content
        content = await file.read()
        audit_service = AuditService(db)

        # Parse based on file type
        if file.filename.endswith('.csv'):
            df = pd.read_csv(
                io.StringIO(content.decode('utf-8')),
                sep=None,
                engine='python'
            )
        else:
            # Read Excel, skip the note row (row 2)
            df = pd.read_excel(
                io.BytesIO(content),
                header=0,  # Headers are in row 1 (index 0)
                skiprows=[1]  # Skip row 2 (index 1) which contains notes
            )

        # Clean column names
        df.columns = df.columns.str.strip()

        # Column mapping to match new template - ADDED Stock Minimum
        column_mapping = {
            'Type': 'type',
            'Nama Item': 'name',
            'SKU': 'sku',
            'Brand': 'brand',
            'Jenis Barang': 'jenis_barang',
            'Jumlah Unit': 'jumlah_unit',
            'Stock Minimum': 'stock_minimum',  # NEW MAPPING
            'Harga Modal': 'harga_modal',
            'Harga Jual': 'harga_jual',
            'Satuan Unit': 'satuan_unit',
            'Nama Vendor': 'nama_vendor'
        }

        # Required columns
        required_columns = ['Nama Item', 'SKU', 'Satuan Unit', 'Harga Jual']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {', '.join(missing_columns)}"
            )

        # Rename columns
        df = df.rename(columns=column_mapping)

        # Build lookup dictionaries
        vendors_lookup = _build_vendors_lookup(db)
        categories_lookup = _build_categories_lookup(db)
        satuans_lookup = _build_satuans_lookup(db)
        existing_skus = _get_existing_skus(db)

        # Initialize result
        result = ImportResult(
            total_processed=0,
            successful_imports=0,
            failed_imports=0,
            errors=[],
            warnings=[]
        )

        # Filter out empty rows (where name and sku are both empty)
        df_filtered = df[
            ~(df['name'].isna() & df['sku'].isna())
        ].copy()

        result.total_processed = len(df_filtered)

        # Process each row
        for index, row in df_filtered.iterrows():
            try:
                item_data = _process_row(
                    row, index, categories_lookup, satuans_lookup,
                    existing_skus, default_item_type, update_existing, vendors_lookup
                )

                if item_data is None:
                    continue

                # Check if item exists
                existing_item = db.query(Item).filter(
                    Item.sku == item_data['sku'],
                    Item.deleted_at.is_(None)
                ).first()

                if update_existing and existing_item:
                    _update_existing_item(
                        db, item_data, existing_item.id,
                        audit_service, user_name
                    )
                else:
                    # Generate unique code
                    prefix = get_item_prefix(item_data['type'])
                    item_code = generate_unique_record_code(db, Item, prefix)
                    item_data['code'] = item_code

                    # Create new item
                    new_item = _create_new_item(
                        db, item_data, audit_service, user_name
                    )
                    existing_skus[item_data['sku']] = new_item.id

                result.successful_imports += 1

            except Exception as e:
                error_msg = str(e)
                result.errors.append({
                    'row': index + 3,  # +3 because: +1 for header, +1 for note row, +1 for 1-based indexing
                    'sku': row.get('sku', 'N/A'),
                    'error': error_msg
                })
                result.failed_imports += 1

                if not skip_on_error:
                    db.rollback()
                    raise HTTPException(
                        status_code=400,
                        detail=f"Row {index + 3}: {error_msg}"
                    )

        # Commit all changes
        db.commit()

        return result

    except pd.errors.EmptyDataError:
        raise HTTPException(
            status_code=400,
            detail="File is empty or has no data"
        )
    except pd.errors.ParserError as e:
        raise HTTPException(
            status_code=400,
            detail=f"Error parsing file: {str(e)}"
        )
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"{str(e)}"
        )


def _get_existing_skus(db: Session) -> Dict[str, int]:
    """Get existing SKUs to check for duplicates."""
    items = db.query(Item.sku, Item.id).filter(
        Item.deleted_at.is_(None)
    ).all()
    return {item.sku: item.id for item in items}


def _process_row(
        row,
        index: int,
        categories_lookup: Dict[str, int],
        satuans_lookup: Dict[str, int],
        existing_skus: Dict[str, int],
        default_item_type: ItemTypeEnum,
        update_existing: bool,
        vendors_lookup: Dict[str, int],
) -> Optional[Dict[str, Any]]:
    """Process a single row from the Excel file."""

    # Validate required fields
    if pd.isna(row.get('name')) or not str(row.get('name')).strip():
        raise ValueError("Nama Item is required")

    if pd.isna(row.get('sku')) or not str(row.get('sku')).strip():
        raise ValueError("SKU is required")

    if pd.isna(row.get('satuan_unit')) or not str(row.get('satuan_unit')).strip():
        raise ValueError("Satuan Unit is required")

    sku = str(row['sku']).strip()

    # Check for duplicate SKU
    if sku in existing_skus and not update_existing:
        raise ValueError(f"SKU '{sku}' already exists.")

    # Validate Satuan Unit
    satuan_symbol = str(row['satuan_unit']).lower().strip()
    satuan_id = satuans_lookup.get(satuan_symbol)
    if not satuan_id:
        raise ValueError(
            f"Satuan '{row['satuan_unit']}' tidak ditemukan. "
            f"Tambahkan entri terlebih dahulu."
        )

    # Process Brand (Category One)
    category_one_id = None
    if not pd.isna(row.get('brand')) and str(row.get('brand')).strip():
        cat1_name = str(row['brand']).lower().strip()
        category_one_id = categories_lookup.get(cat1_name)
        if not category_one_id:
            raise ValueError(
                f"Brand '{row['brand']}' tidak ditemukan. "
                f"Tambahkan entri terlebih dahulu."
            )

    # Process Jenis Barang (Category Two)
    category_two_id = None
    if not pd.isna(row.get('jenis_barang')) and str(row.get('jenis_barang')).strip():
        cat2_name = str(row['jenis_barang']).lower().strip()
        category_two_id = categories_lookup.get(cat2_name)
        if not category_two_id:
            raise ValueError(
                f"Jenis Barang '{row['jenis_barang']}' tidak ditemukan. "
                f"Tambahkan entri terlebih dahulu."
            )
            
    vendor_id = None
    if not pd.isna(row.get('nama_vendor')) and str(row.get('nama_vendor')).strip():
        vendor_name = str(row['nama_vendor']).lower().strip()
        vendor_id = vendors_lookup.get(vendor_name)
        if not vendor_id:
            raise ValueError(
                f"Vendor '{row['nama_vendor']}' tidak ditemukan. "
                f"Tambahkan entri terlebih dahulu."
            )

    # Process Harga Modal (Cost Price)
    try:
        if pd.isna(row.get('harga_modal')) or str(row.get('harga_modal')).strip() == '':
            modal_price = Decimal('0')
        else:
            # Handle both comma and dot as decimal separators
            cost_value = str(row['harga_modal']).replace(',', '.')
            modal_price = Decimal(cost_value)
            if modal_price < 0:
                raise ValueError("Harga Modal must be positive")
    except (ValueError, TypeError) as e:
        raise ValueError(f"Invalid Harga Modal: {row.get('harga_modal', 'N/A')}")

    # Process Harga Jual (Selling Price)
    try:
        if pd.isna(row.get('harga_jual')) or str(row.get('harga_jual')).strip() == '':
            selling_price = Decimal('0')
        else:
            # Handle both comma and dot as decimal separators
            sell_value = str(row['harga_jual']).replace(',', '.')
            selling_price = Decimal(sell_value)
            if selling_price < 0:
                raise ValueError("Harga Jual must be positive")
    except (ValueError, TypeError) as e:
        raise ValueError(f"Invalid Harga Jual: {row.get('harga_jual', 'N/A')}")

    # Process Jumlah Unit (Quantity)
    total_item = 0
    if not pd.isna(row.get('jumlah_unit')):
        try:
            total_item = int(float(row['jumlah_unit']))
            if total_item < 0:
                raise ValueError("Jumlah Unit must be non-negative")
        except (ValueError, TypeError):
            raise ValueError(f"Invalid Jumlah Unit: {row.get('jumlah_unit', 'N/A')}")

    # Process Stock Minimum (NEW)
    min_item = 0
    if not pd.isna(row.get('stock_minimum')):
        try:
            min_item = int(float(row['stock_minimum']))
            if min_item < 0:
                raise ValueError("Stock Minimum must be non-negative")
        except (ValueError, TypeError):
            raise ValueError(f"Invalid Stock Minimum: {row.get('stock_minimum', 'N/A')}")

    # Process Type column
    type_value = str(row.get('type', '')).strip().lower() if not pd.isna(row.get('type')) else None
    type_mapping = {
        "high quality": ItemTypeEnum.HIGH_QUALITY,
        "raw material": ItemTypeEnum.RAW_MATERIAL,
        "service": ItemTypeEnum.SERVICE,
    }

    if type_value and type_value in type_mapping:
        item_type = type_mapping[type_value]
    else:
        item_type = default_item_type

    return {
        'name': str(row['name']).strip(),
        'sku': sku,
        'type': item_type,
        'total_item': total_item,
        'min_item': min_item,  # NEW FIELD
        'modal_price': modal_price,  # Harga Modal
        'price': selling_price,    # Harga Jual
        'category_one': category_one_id,
        'category_two': category_two_id,
        'satuan_id': satuan_id,
        'vendor_id': vendor_id,
        'is_active': True
    }

    
@router.put("/{item_id}", response_model=ItemResponse)
async def update_item(
        request: Request,
        item_id: int,
        type: ItemTypeEnum = Form(...),
        name: str = Form(...),
        sku: str = Form(...),
        total_item: int = Form(0),
        min_item: int = Form(0),
        price: float = Form(...),
        modal_price: float = Form(...),
        is_active: bool = Form(True),
        category_one: Optional[int] = Form(None),
        category_two: Optional[int] = Form(None),
        vendor_id: Optional[str] = Form(None),
        satuan_id: int = Form(...),
        images: List[UploadFile] = File(default=[]),
        existing_images: List[int] = Form(default=[]),  # IDs of images to keep
        deleted_images: List[int] = Form(default=[]),   # IDs of images to delete
        db: Session = Depends(get_db),
        user_name: str = Depends(get_current_user_name),
):
    # Validate file sizes before processing
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB per file
    MAX_TOTAL_SIZE = 30 * 1024 * 1024  # 30MB total

    audit_service = AuditService(db)

    total_size = 0
    for image in images:
        if image.filename:
            image.file.seek(0, 2)
            file_size = image.file.tell()
            image.file.seek(0)

            if file_size > MAX_FILE_SIZE:
                raise HTTPException(
                    status_code=413,
                    detail=f"File {image.filename} is too large. Maximum size is 10MB per file."
                )

            total_size += file_size

    if total_size > MAX_TOTAL_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"Total file size is too large. Maximum total size is 30MB."
        )

    db_item = db.query(Item).filter(Item.id == item_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Item tidak ditemukan")

    existing_item = db.query(Item).filter(Item.sku == sku, Item.id != item_id).first()
    if existing_item:
        raise HTTPException(status_code=400, detail="SKU sudah ada")

    try:
        # Update basic fields
        db_item.type = type
        db_item.name = name
        db_item.sku = sku
        db_item.total_item = total_item
        db_item.min_item = min_item
        db_item.price = price
        db_item.modal_price = modal_price
        db_item.is_active = is_active
        db_item.category_one = category_one
        db_item.category_two = category_two
        db_item.vendor_id = vendor_id
        db_item.satuan_id = satuan_id

        # Handle image deletions
        if deleted_images:
            attachments_to_delete = db.query(AllAttachment).filter(
                AllAttachment.item_id == item_id,
                AllAttachment.id.in_(deleted_images)
            ).all()

            for attachment in attachments_to_delete:
                # Delete physical file
                if os.path.exists(attachment.file_path):
                    os.remove(attachment.file_path)
                # Delete database record
                db.delete(attachment)

        # Get current attachment count (after deletions)
        current_attachment_count = db.query(AllAttachment).filter(
            AllAttachment.item_id == item_id
        ).count()

        # Add new images
        new_image_count = len([img for img in images if img.filename])
        total_images = current_attachment_count + new_image_count

        if total_images > 3:
            raise HTTPException(
                status_code=400,
                detail=f"Maximum 3 images allowed. Currently have {current_attachment_count}, trying to add {new_image_count}"
            )

        for image in images:
            if image.filename:
                # Validate file type
                allowed_types = {'image/jpeg', 'image/png', 'image/gif', 'image/webp'}
                if image.content_type not in allowed_types:
                    raise HTTPException(
                        status_code=400,
                        detail=f"File type {image.content_type} not allowed. Allowed types: JPEG, PNG, GIF, WebP"
                    )

                ext = os.path.splitext(image.filename)[1]
                unique_filename = f"{uuid.uuid4()}{ext}"
                save_path = os.path.join(NEXT_PUBLIC_UPLOAD_DIR, unique_filename)

                with open(save_path, "wb") as buffer:
                    shutil.copyfileobj(image.file, buffer)

                attachment = AllAttachment(
                    parent_type=ParentType.ITEMS,
                    item_id=db_item.id,
                    filename=image.filename,
                    file_path=save_path,
                    file_size=os.path.getsize(save_path),
                    mime_type=image.content_type,
                    created_at=datetime.now(),
                )
                db.add(attachment)

        db.commit()
        db.refresh(db_item)

        audit_service.default_log(
            entity_id=db_item.id,
            entity_type=AuditEntityEnum.ITEM,
            description=f"Item {db_item.name} telah diubah",
            user_name=user_name,
        )

        return construct_item_response(db_item, request)

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating item: {str(e)}")

@router.delete("/{item_id}")
def delete_item(item_id: int, db: Session = Depends(get_db)):
    db_item = db.query(Item).filter(Item.id == item_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Item tidak ditemukan")

    try:
        for attachment in db_item.attachments:
            if os.path.exists(attachment.file_path):
                os.remove(attachment.file_path)
            db.delete(attachment)

        soft_delete_record(db, Item, item_id)

        db.commit()
        return {"message": f"Item {item_id} soft deleted successfully (attachments removed)"}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting item: {str(e)}")


def construct_item_response(item: Item, request: Request) -> Dict[str, Any]:
    static_url = os.environ.get("BASE_URL", "http://localhost:8000/static")

    enriched_attachments = []
    for att in item.attachments:
        clean_path = att.file_path.replace("\\", "/").replace("uploads/", "")
        enriched_attachments.append({
            "id": att.id,
            "filename": att.filename,
            "file_path": att.file_path,
            "file_size": att.file_size,
            "mime_type": att.mime_type,
            "created_at": att.created_at,
            "url": f"{static_url}/{clean_path}",
        })

    return {
        "id": item.id,
        "type": item.type,
        "name": item.name,
        "sku": item.sku,
        "code": item.code,

        "min_item": item.min_item,
        "modal_price": item.modal_price,
        "total_item": item.total_item,
        "price": item.price,
        "is_active": item.is_active,

        "created_at": getattr(item, "created_at", None),
        "category_one_rel": CategoryOut.model_validate(item.category_one_rel).model_dump() if item.category_one_rel else None,
        "category_two_rel": CategoryOut.model_validate(item.category_two_rel).model_dump() if item.category_two_rel else None,
        "satuan_rel": SatuanOut.model_validate(item.satuan_rel).model_dump() if item.satuan_rel else None,
        "vendor_rel": VendorOut.model_validate(item.vendor_rel).model_dump() if item.vendor_rel else None,
        "attachments": enriched_attachments,
    }
